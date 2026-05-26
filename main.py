import os
import openpyxl

from playwright.sync_api import Page, sync_playwright
from dotenv import load_dotenv
from time import sleep
from datetime import date, datetime
from pandas import pandas as pd, DataFrame
from tasy_automation_helper import Logar, AbrirFuncao, ObterPainel, Formulario as form, Operadores

MAIN_URL = "https://tasycorp-vivere.zion-srv.com/#/login"

def formatar_data(data:str, format:str='%d/%m/%Y %H:%M:%S')->str:    
    return datetime.strptime(data,format).strftime('%d/%m/%Y %H:%M:%S')

def fechar_modal_operacao_abortada(page:Page) -> Page:
    selector = 'div[role="alertdialog"].priority-dialog'
    modal = page.locator(selector)
    modal.locator('button#w-dialog-box-ok-button').click()
    return page

def fechar_modal_pr_atividade(page:Page) -> Page:
    selector = 'div[role="alertdialog"] [w-code="1099967"]'
    modal = page.locator(selector)
    modal.locator('button.gwt-Button.btn-gray').click()
    
    return page

def aguardar_carregamento(page:Page, locator:str='div.paginationtop'):
    '''
    Aguardar a tela carregar, baseado no locator (locator). Padrão: div.paginationtop
    '''
    page.wait_for_selector(locator)
    pass

def filtrar_projeto(page:Page, nr_seq_projeto:int) -> Page:
    '''
    Filtra e abre um projeto. Recebe a página atual e o número do projeto a ser aberto.
    '''
    projeto = nr_seq_projeto    
    aguardar_carregamento(page=page)

    # Abre a tela do filtro
    sleep(3)
    page.wait_for_selector("div.token-filter-container")    
    icone_filtro = page.locator("div.token-filter-container")
    icone_filtro = icone_filtro.locator("tasy-wlabel.filter-icon.filter-icon-blue").nth(0)
    if icone_filtro:
        icone_filtro.click()
    

    # Seleciona apenas o painel do filtro
    page.wait_for_selector("div.wfilter")
    painel_filtro = page.locator("div.wfilter")
    
    
    # Preenche o campo sequencia projeto
    campo_sequencia = painel_filtro.locator('[data-testid="NR_SEQUENCIA"] input[name="NR_SEQUENCIA"]')
    campo_sequencia.fill(str(projeto))

    # Clicar no botão filtrar
    painel_filtro.locator('button.btn-green.wfilter-button.ng-binding').click()

    return page

def filtrar_rat(page:Page, nr_seq_rat:int) -> Page:
    aguardar_carregamento(page=page, locator='[w-code="1094205"]')
    painel = page.locator('[w-code="1094205"]')
    icone_filtro = painel.locator("div.token-filter-container tasy-wlabel.filter-icon.filter-icon-blue")    
    icone_filtro.click()

    painel_filtro = page.locator('[w-code="1094278"] tasy-wfilter')
    campo_sequencia = painel_filtro.locator('[data-testid="NR_SEQUENCIA"] input[name="NR_SEQUENCIA"]')

    campo_sequencia.fill(str(nr_seq_rat))

    # Clicar no botão filtrar
    painel_filtro.locator('button.btn-green.wfilter-button.ng-binding').click()

    return page

def abrir_projeto(page:Page, coluna:int=0) -> Page:
    '''
    Clica na primeira linha do painel projetos
    '''
    # data-row-idx="0"
    # wdbpanel-container
    page = page
    # page.wait_for_timeout(9999)
    page.wait_for_selector('div#layout')    
    painel = page.locator('div#layout div.wdbpanel-container')
    row = painel.locator('[data-row-idx="0"][role="presentation"]')
    row.dblclick()
    # col=coluna
    return page

def abrir_rat(page:Page, coluna:int=0) -> Page:
    '''
    Clica na primeira linha do painel de RAT
    '''
    aguardar_carregamento(page, '[w-code="1094278"]')
    painel = page.locator('[w-code="1094278"]') #div#layout div.wdbpanel-container')
    linha = painel.locator( 'div#layout div.wdbpanel-container [data-row-idx="0"][role="row"]' )
    linha.dblclick()
    
    return page

def preecher_form_atividade(page:Page, wdbpanel:int|str, dados: tuple) -> Page:
    painel = ObterPainel.obter_painel(page, wdbpanel)
    dt_inicio = formatar_data( str(dados.DT_INICIO), '%Y-%m-%d %H:%M:%S' )
    dt_fim = formatar_data( str(dados.DT_FIM), '%Y-%m-%d %H:%M:%S' )
    atividade_id = dados.ATIVIDADE_ID
    classificacao = dados.CLASSIFICACAO
    atividade = dados.DESCRICAO
    modalidade = dados.MODALIDADE

    form.preecher_data(painel=painel, atributo="DT_INICIO_ATIV", data=dt_inicio)
    form.preecher_data(painel=painel, atributo="DT_FIM_ATIV", data= dt_fim)
    form.preecher_numero(painel=painel, atributo="NR_SEQ_ETAPA_CRON",num=atividade_id )
    form.preecher_lookup(painel=painel, atributo="IE_CLASSIFICACAO",texto=classificacao)
    form.preecher_texto(painel=painel, atributo="ds_atividade",texto=atividade )
    form.preencher_radiobutton(painel=painel, atributo="IE_MODALIDADE", valor=modalidade )
    return page

def marcar_linha_como_lancada(index: int, status: str = "Sim"):
    """Abre o arquivo Excel e grava o status na linha correspondente."""
    wb = openpyxl.load_workbook(ARQUIVO)
    ws = wb.active

    # Verificar se a coluna CONTROLE já existe, senão cria
    headers = [cell.value for cell in ws[1]]
    if "LANCADA" not in headers:
        ie_lancada = len(headers) + 1
        ws.cell(row=1, column=ie_lancada, value="LANCADA")
    else:
        ie_lancada = headers.index("LANCADA") + 1

    # index do itertuples começa em 0, mas no Excel linha 1 = header, linha 2 = primeira linha de dados
    row_excel = index + 2  # +1 pelo header, +1 porque index pandas começa em 0

    ws.cell(row=row_excel, column=ie_lancada, value=status)
    wb.save(ARQUIVO)



def carregar_arquivo_dados():
    df = pd.read_excel(ARQUIVO)
    df = df.reset_index(drop=True)
    return df


def lancar_rats(page:Page, dados:DataFrame):
    df = dados

    df = dados[dados["LANCADA"] != "Sim"] if "LANCADA" in dados.columns else dados
    qtde_rats = len(df)
    if qtde_rats == 0:
        print("Nada há lançar")
        return

    print(f"Qtde total de RATs a serem Lançadas: {qtde_rats}")

    projetos = df["PROJETO_ID"].unique()

    for projeto in projetos:
        # Filtrar projetos
        page = filtrar_projeto(page=page, nr_seq_projeto=projeto)
        page = abrir_projeto(page=page)
        rats = df[df["PROJETO_ID"]  == projeto]["RAT_ID"].unique()
        for rat in rats:
            # Filtrar RAT
            page = filtrar_rat(page=page, nr_seq_rat=rat)
            # Abrir RAT
            page = abrir_rat(page=page)        
            atividades = df[ (df["PROJETO_ID"] == projeto) & (df["RAT_ID"] == rat) ]
            indice = 0
            for atividade in atividades.itertuples():                
                indice+=1
                page = Operadores.adicionar(page=page, wdbpanel=1094287)
                page = preecher_form_atividade(page=page, wdbpanel=109287,dados=atividade)

                sleep(1)
                page = Operadores.salvar(page=page, wdbpanel=1094287)
                sleep(2)
                page = fechar_modal_operacao_abortada(page=page)
                sleep(2)
                page = fechar_modal_pr_atividade(page=page)
                marcar_linha_como_lancada(index=atividade.Index)
                print(f'RAT: {atividade.RAT_ID} - {indice} de { len(atividades)} lançadas.')

            
        
def run():
    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=False,channel="chrome",
            args=[
                "--start-maximized",
                "--disable-notifications",
                "--disable-features=ExternalProtocolDialog"
            ]
        )

        context = browser.new_context(
            permissions=[],
            no_viewport=True  # permite usar o tamanho real da janela maximizada
        )
        page = context.new_page()
        page.goto(MAIN_URL)        

        # Login
        usr = os.getenv("TASY_USR")
        pwd = os.getenv("TASY_PWD")
        page = Logar.realizar_login(page=page, usuario=usr, senha=pwd)
        
        # Abrir Função
        page = AbrirFuncao.abrir_funcao(page=page, nome_funcao="Gestão de Projetos Philips")
        
        df = carregar_arquivo_dados()
        lancar_rats(page, df)
        
        page.wait_for_timeout(5000)
        print('Fim do processo')
        browser.close()


if __name__ == "__main__":
    ARQUIVO = r'.\data\dados.xlsx'
    load_dotenv()    
    run()